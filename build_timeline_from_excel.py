#!/usr/bin/env python3
"""
Build timeline-data.js from excel/armenian_history_timeline.xlsx.

Data sheets: names that do NOT end with _translation or _translation_desc.

Translation sheets (joined on * *_id matching data sheet):
  Prefer {category}_translation with columns such as:
    history_id | history_section_en | history_description_en | history_section_ru | ...
  Farsi columns use suffix _ir (mapped to locale "fa").

Legacy {category}_translation_desc with only *_description_{en,ru,ir,fr} is still supported.

Requires: pip install openpyxl

Optional: python build_timeline_from_excel.py --sync-translation-rows
  Appends missing id rows to each {category}_translation sheet (empty translation cells).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "excel" / "armenian_history_timeline.xlsx"
OUT = ROOT / "timeline-data.js"

COSMO_THRESHOLD = 271_821

# UI + timeline styling
CATEGORY_META: dict[str, dict] = {
    "hist": {
        "color": "#1f4e79",
        "labels": {
            "hy": "Պատմություն",
            "en": "History",
            "ru": "История",
            "fa": "تاریخ",
            "fr": "Histoire",
        },
    },
    "leaders": {
        "color": "#6c3483",
        "labels": {
            "hy": "Առաջնորդներ",
            "en": "Leaders",
            "ru": "Лидеры",
            "fa": "رهبران",
            "fr": "Dirigeants",
        },
    },
    "flag": {
        "color": "#c0392b",
        "labels": {
            "hy": "Դրոշ",
            "en": "Flag",
            "ru": "Флаг",
            "fa": "پرچم",
            "fr": "Drapeau",
        },
    },
    "map": {
        "color": "#117a65",
        "labels": {
            "hy": "Քարտեզ",
            "en": "Map",
            "ru": "Карта",
            "fa": "نقشه",
            "fr": "Carte",
        },
    },
    "population": {
        "color": "#b7950b",
        "labels": {
            "hy": "Բնակչություն",
            "en": "Population",
            "ru": "Население",
            "fa": "جمعیت",
            "fr": "Population",
        },
    },
    "gdp": {
        "color": "#2471a3",
        "labels": {
            "hy": "Տնտեսություն",
            "en": "Economy",
            "ru": "Экономика",
            "fa": "اقتصاد",
            "fr": "Économie",
        },
    },
    "stem": {
        "color": "#34495e",
        "labels": {
            "hy": "Գիտություն և տեխնոլոգիա",
            "en": "STEM",
            "ru": "Наука и технологии",
            "fa": "علم و فناوری",
            "fr": "Science et technologie",
        },
    },
    "sports": {
        "color": "#ca6f1e",
        "labels": {
            "hy": "Սպորտ",
            "en": "Sports",
            "ru": "Спорт",
            "fa": "ورزش",
            "fr": "Sports",
        },
    },
    "music": {
        "color": "#af7ac5",
        "labels": {
            "hy": "Երաժշտություն",
            "en": "Music",
            "ru": "Музыка",
            "fa": "موسیقی",
            "fr": "Musique",
        },
    },
    "literature": {
        "color": "#148f77",
        "labels": {
            "hy": "Գրականություն",
            "en": "Literature",
            "ru": "Литература",
            "fa": "ادبیات",
            "fr": "Littérature",
        },
    },
}

# Excel column suffix -> bundle locale key (Farsi uses _ir in sheets)
LOCALE_BY_SUFFIX: dict[str, str] = {
    "en": "en",
    "ru": "ru",
    "ir": "fa",
    "fr": "fr",
}


def _is_translation_sheet(name: str) -> bool:
    n = name.lower()
    return n.endswith("_translation") or n.endswith("_translation_desc")


def _is_data_sheet(name: str) -> bool:
    return not _is_translation_sheet(name)


def _translation_sheet_for_category(wb: openpyxl.Workbook, category: str) -> str | None:
    """Pick best matching translation sheet (prefer *_translation over *_translation_desc)."""
    candidates: list[str] = []
    if category == "leaders":
        candidates.extend(["leaders_translation", "people_translation"])
    candidates.append(f"{category}_translation")
    if category == "leaders":
        candidates.append("people_translation_desc")
    candidates.append(f"{category}_translation_desc")

    seen: set[str] = set()
    for name in candidates:
        if name in seen:
            continue
        seen.add(name)
        if name in wb.sheetnames:
            return name
    return None


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _headline_from_text(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return "—"
    first = t.split("\n")[0].strip()
    first = re.sub(r"\s+", " ", first)
    if len(first) > 140:
        return first[:137].rstrip() + "…"
    return first


def _desc_html(description: str, rtl: bool = False) -> str:
    if not (description or "").strip():
        return ""
    parts = [p.strip() for p in description.split("\n") if p.strip()]
    if not parts:
        return ""
    inner = "".join(f"<p>{_escape_html(p)}</p>" for p in parts)
    if rtl:
        return f'<div dir="rtl" lang="fa">{inner}</div>'
    return inner


def _escape_html(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _year_pair(section_id: str, start, finish) -> tuple[int, int] | None:
    if start is None and finish is None:
        return None
    if start is None:
        start = finish
    if finish is None:
        finish = start
    try:
        y0 = int(float(start))
        y1 = int(float(finish))
    except (TypeError, ValueError):
        return None
    sid = (section_id or "").strip().upper()
    if sid == "BCE":
        y0_j = -abs(y0)
        y1_j = -abs(y1)
    elif sid in ("CE", "AD", "ԹՎ", ""):
        y0_j = abs(y0)
        y1_j = abs(y1)
    else:
        raise ValueError(f"Unknown section_id: {section_id!r}")
    return (min(y0_j, y1_j), max(y0_j, y1_j))


def _display_date(section_id: str, start, finish) -> str:
    sid = (section_id or "").strip().upper()
    try:
        s = int(float(start)) if start is not None else int(float(finish))
        f = int(float(finish)) if finish is not None else int(float(start))
    except (TypeError, ValueError):
        return ""
    if sid == "BCE":
        if s == f:
            return f"մ.թ.ա. {s:,}"
        older, newer = max(s, f), min(s, f)
        return f"մ.թ.ա. {older:,} — {newer:,}"
    if s == f:
        return f"{s} թ."
    a, b = min(s, f), max(s, f)
    return f"{a}—{b} թ."


def _row_dict(header: list[str], row: tuple) -> dict[str, object]:
    out: dict[str, object] = {}
    for i, key in enumerate(header):
        if key:
            out[str(key).strip()] = row[i] if i < len(row) else None
    return out


def _get_by_suffix(d: dict[str, object], suffix: str) -> object | None:
    for k, v in d.items():
        if k.lower().endswith(suffix.lower()):
            return v
    return None


def _parse_translation_header(header: list[str]) -> tuple[str | None, dict[str, tuple[str | None, str | None]]]:
    """
    Returns (id_column_name, locale -> (section_col, description_col)).
    Matches columns ending with _section_{en,ru,ir,fr} and _description_{en,ru,ir,fr}.
    """
    id_col: str | None = None
    for h in header:
        if not h:
            continue
        hl = h.lower()
        if hl.endswith("_id") and "translation" not in hl:
            id_col = h
            break
    if id_col is None and header:
        id_col = header[0]

    pairs: dict[str, tuple[str | None, str | None]] = {
        loc: (None, None) for loc in ("en", "ru", "fa", "fr")
    }

    for h in header:
        if not h:
            continue
        hl = h.lower()
        for suf, loc in LOCALE_BY_SUFFIX.items():
            if hl.endswith(f"_section_{suf}"):
                sec, desc = pairs[loc]
                pairs[loc] = (h, desc)
            elif hl.endswith(f"_description_{suf}"):
                sec, desc = pairs[loc]
                pairs[loc] = (sec, h)

    return id_col, pairs


def _load_translations(wb: openpyxl.Workbook, category: str) -> dict[str, dict[str, dict[str, str]]]:
    """
    row_id -> locale -> {"section": str, "description": str}
    """
    tname = _translation_sheet_for_category(wb, category)
    if not tname:
        return {}

    ws = wb[tname]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    header = [_cell_str(c) for c in rows[0]]
    id_col, col_pairs = _parse_translation_header(header)
    idx = {name: i for i, name in enumerate(header) if name}

    out: dict[str, dict[str, dict[str, str]]] = {}

    for row in rows[1:]:
        if not row or all(c is None or _cell_str(c) == "" for c in row):
            continue
        rd = _row_dict(header, row)
        rid = _cell_str(rd.get(id_col or "", row[0] if row else ""))
        if not rid:
            continue

        per_loc: dict[str, dict[str, str]] = {}

        for loc, (sec_h, desc_h) in col_pairs.items():
            sec = ""
            desc = ""
            if sec_h and sec_h in idx:
                i = idx[sec_h]
                if i < len(row) and row[i] is not None:
                    sec = str(row[i]).strip()
            if desc_h and desc_h in idx:
                i = idx[desc_h]
                if i < len(row) and row[i] is not None:
                    desc = str(row[i]).strip()

            if sec or desc:
                per_loc[loc] = {"section": sec, "description": desc}

        # Legacy: only description columns (no *_section_*)
        if not per_loc:
            for suf, loc in LOCALE_BY_SUFFIX.items():
                legacy = None
                for h in header:
                    hl = h.lower()
                    if hl == f"history_description_{suf}" or hl.endswith(f"_description_{suf}"):
                        if "section" not in hl:
                            legacy = h
                            break
                if legacy and legacy in idx:
                    i = idx[legacy]
                    if i < len(row) and row[i] is not None:
                        d = str(row[i]).strip()
                        if d:
                            per_loc[loc] = {"section": "", "description": d}

        if per_loc:
            out[rid] = per_loc

    return out


def _needs_cosmological(years: list[int]) -> bool:
    return any(abs(y) > COSMO_THRESHOLD for y in years)


def _locale_entry(
    loc: str,
    section_hy: str,
    hy_headline: str,
    hy_html: str,
    tr: dict[str, dict[str, str]] | None,
) -> tuple[str, str]:
    """Returns (headline, html) for one locale."""
    t = (tr or {}).get(loc)
    if not t:
        return hy_headline, hy_html

    sec = (t.get("section") or "").strip()
    desc = (t.get("description") or "").strip()

    if desc:
        body = _desc_html(desc, rtl=(loc == "fa"))
    else:
        body = hy_html

    if sec:
        headline = _headline_from_text(sec)
    elif desc:
        headline = _headline_from_text(desc)
    else:
        headline = hy_headline

    return headline, body


def _group_labels(
    section_hy: str,
    tr: dict[str, dict[str, str]] | None,
) -> dict[str, str]:
    g: dict[str, str] = {}
    if section_hy:
        g["hy"] = section_hy
    if not tr:
        for loc in ("en", "ru", "fa", "fr"):
            if section_hy:
                g[loc] = section_hy
        return g

    for loc in ("en", "ru", "fa", "fr"):
        t = tr.get(loc)
        sec = (t.get("section") or "").strip() if t else ""
        if sec:
            g[loc] = sec
        elif section_hy:
            g[loc] = section_hy
    return g


def _parse_data_sheet(ws, category: str, color: str, translations: dict[str, dict[str, dict[str, str]]]):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    header = [_cell_str(c) for c in rows[0]]
    events = []
    all_years: list[int] = []

    for row in rows[1:]:
        if not row or all(c is None or _cell_str(c) == "" for c in row):
            continue
        rd = _row_dict(header, row)
        rid = _cell_str(_get_by_suffix(rd, "_id") or rd.get(header[0] if header else "") or row[0])
        if not rid:
            continue
        section_title = _cell_str(_get_by_suffix(rd, "_section_title") or "")
        section_id = _cell_str(_get_by_suffix(rd, "_section_id") or "")
        start = _get_by_suffix(rd, "_start")
        finish = _get_by_suffix(rd, "_finish")
        desc_raw = _get_by_suffix(rd, "_description")
        if desc_raw is None:
            desc_raw = ""
        elif isinstance(desc_raw, float) and desc_raw.is_integer():
            desc_raw = str(int(desc_raw))
        else:
            desc_raw = str(desc_raw)

        pair = _year_pair(section_id, start, finish)
        if pair is None:
            continue
        y_start, y_end = pair
        all_years.extend([y_start, y_end])

        hy_headline = _headline_from_text(desc_raw) if desc_raw else _headline_from_text(section_title)
        hy_html = _desc_html(desc_raw, rtl=False)

        tr_row = translations.get(rid, {})

        locales: dict[str, dict[str, str]] = {
            "hy": {"headline": hy_headline, "text": hy_html},
        }
        for loc in ("en", "ru", "fa", "fr"):
            h, html = _locale_entry(loc, section_title, hy_headline, hy_html, tr_row)
            locales[loc] = {"headline": h, "text": html}

        group = _group_labels(section_title, tr_row)

        dd = _display_date(section_id, start, finish)
        ev: dict = {
            "category": category,
            "unique_id": f"{category}::{rid}",
            "start_date": {"year": y_start},
            "locales": locales,
            "group": group,
            "background": {"color": color},
        }
        if dd:
            ev["start_date"]["display_date"] = dd
        if y_end != y_start:
            ev["end_date"] = {"year": y_end}

        events.append(ev)

    return events, all_years


def _collect_ids_from_data_sheet(ws) -> list[str]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [_cell_str(c) for c in rows[0]]
    ids: list[str] = []
    for row in rows[1:]:
        if not row:
            continue
        rd = _row_dict(header, row)
        rid = _cell_str(_get_by_suffix(rd, "_id") or rd.get(header[0] if header else "") or row[0])
        if rid:
            ids.append(rid)
    return ids


def sync_translation_row_templates() -> int:
    """Ensure each {category}_translation sheet has a row per data id (empty cells)."""
    if not XLSX.is_file():
        print(f"Missing workbook: {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=False)

    for category in list(CATEGORY_META.keys()):
        tname = f"{category}_translation"
        if tname not in wb.sheetnames:
            continue
        if category not in wb.sheetnames:
            continue

        ws_d = wb[category]
        ws_t = wb[tname]
        want_ids = _collect_ids_from_data_sheet(ws_d)

        header = [_cell_str(c) for c in next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        if not header:
            continue
        existing = set()
        for row in ws_t.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                existing.add(_cell_str(row[0]))

        n_added = 0
        for rid in want_ids:
            if rid in existing:
                continue
            ws_t.append([rid] + [None] * (len(header) - 1))
            existing.add(rid)
            n_added += 1

        if n_added:
            print(f"{tname}: added {n_added} row(s).", file=sys.stderr)

    wb.save(XLSX)
    wb.close()
    print(f"Saved {XLSX}")
    return 0


def build_bundle() -> int:
    if not XLSX.is_file():
        print(f"Missing workbook: {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    data_sheet_names = [s for s in wb.sheetnames if _is_data_sheet(s)]

    order_pref = list(CATEGORY_META.keys())
    data_sheet_names.sort(
        key=lambda s: (order_pref.index(s) if s in order_pref else 999, s)
    )

    all_events: list[dict] = []
    all_years: list[int] = []
    categories_out: list[dict] = []

    for sheet_name in data_sheet_names:
        meta = CATEGORY_META.get(
            sheet_name,
            {
                "color": "#5d6d7e",
                "labels": {
                    "hy": sheet_name,
                    "en": sheet_name,
                    "ru": sheet_name,
                    "fa": sheet_name,
                    "fr": sheet_name,
                },
            },
        )
        color = meta["color"]
        translations = _load_translations(wb, sheet_name)
        ws = wb[sheet_name]
        evs, ys = _parse_data_sheet(ws, sheet_name, color, translations)
        all_events.extend(evs)
        all_years.extend(ys)
        categories_out.append(
            {
                "id": sheet_name,
                "color": color,
                "labels": meta["labels"],
                "count": len(evs),
            }
        )

    wb.close()

    if not all_events:
        print("No timeline rows found (need numeric start/finish and BCE/CE).", file=sys.stderr)
        return 1

    cosmo = _needs_cosmological(all_years)

    title_strings = {
        "hy": {
            "headline": "Հայաստան — ժամանակագրություն",
            "text": (
                "<p>Ընտրեք բաժինները և լեզուն վերևի վահանակից։ Տվյալները՝ "
                "<code>excel/armenian_history_timeline.xlsx</code>։ Թարմացնել՝ "
                "<code>python build_timeline_from_excel.py</code>։</p>"
                "<p>Թարգմանություններ՝ համապատասխան <code>*_translation</code> թերթերում "
                "(օր. <code>hist_translation</code>)՝ <code>history_id</code>-ով կապված։</p>"
                "<p><a href=\"https://timeline.knightlab.com/\" target=\"_blank\" rel=\"noopener\">TimelineJS</a> "
                "(Knight Lab, MPL 2.0)</p>"
            ),
        },
        "en": {
            "headline": "Armenia — timeline",
            "text": (
                "<p>Choose categories and language in the bar above. Data: "
                "<code>excel/armenian_history_timeline.xlsx</code>. Regenerate with "
                "<code>python build_timeline_from_excel.py</code>.</p>"
                "<p>Translations: matching <code>*_translation</code> sheets (e.g. "
                "<code>hist_translation</code>), keyed by <code>history_id</code>.</p>"
                "<p><a href=\"https://timeline.knightlab.com/\" target=\"_blank\" rel=\"noopener\">TimelineJS</a> "
                "(Knight Lab, MPL 2.0)</p>"
            ),
        },
        "ru": {
            "headline": "Армения — хронология",
            "text": (
                "<p>Выберите категории и язык на панели сверху. Данные: "
                "<code>excel/armenian_history_timeline.xlsx</code>. Обновить: "
                "<code>python build_timeline_from_excel.py</code>.</p>"
                "<p>Переводы: листы <code>*_translation</code> (напр. <code>hist_translation</code>), "
                "связь по <code>history_id</code>.</p>"
                "<p><a href=\"https://timeline.knightlab.com/\" target=\"_blank\" rel=\"noopener\">TimelineJS</a> "
                "(Knight Lab, MPL 2.0)</p>"
            ),
        },
        "fa": {
            "headline": "ارمنستان — خط زمان",
            "text": (
                '<div dir="rtl" lang="fa"><p>دسته‌ها و زبان را از نوار بالا انتخاب کنید. داده‌ها: '
                "<code>excel/armenian_history_timeline.xlsx</code>. برای به‌روزرسانی: "
                "<code>python build_timeline_from_excel.py</code>.</p>"
                "<p>ترجمه‌ها در برگه‌های <code>*_translation</code> (مثلاً "
                "<code>hist_translation</code>) با کلید <code>history_id</code>.</p>"
                "<p><a href=\"https://timeline.knightlab.com/\" target=\"_blank\" rel=\"noopener\">TimelineJS</a> "
                "(Knight Lab, MPL 2.0)</p></div>"
            ),
        },
        "fr": {
            "headline": "Arménie — chronologie",
            "text": (
                "<p>Choisissez les catégories et la langue dans la barre supérieure. Données : "
                "<code>excel/armenian_history_timeline.xlsx</code>. Régénérer avec "
                "<code>python build_timeline_from_excel.py</code>.</p>"
                "<p>Traductions : feuilles <code>*_translation</code> (ex. "
                "<code>hist_translation</code>), liées par <code>history_id</code>.</p>"
                "<p><a href=\"https://timeline.knightlab.com/\" target=\"_blank\" rel=\"noopener\">TimelineJS</a> "
                "(Knight Lab, MPL 2.0)</p>"
            ),
        },
    }

    bundle = {
        "title": title_strings,
        "categories": categories_out,
        "events": all_events,
    }
    if cosmo:
        bundle["scale"] = "cosmological"

    out_text = (
        "/* Auto-generated by build_timeline_from_excel.py */\n"
        "window.TIMELINE_BUNDLE = "
        + json.dumps(bundle, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUT.write_text(out_text, encoding="utf-8")
    print(
        f"Wrote {OUT} ({len(all_events)} events, {len(categories_out)} categories, "
        f"scale={'cosmological' if cosmo else 'human'})"
    )
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Build timeline JSON from Excel.")
    ap.add_argument(
        "--sync-translation-rows",
        action="store_true",
        help="Append missing id rows to each {category}_translation sheet, then exit.",
    )
    args = ap.parse_args()

    if args.sync_translation_rows:
        return sync_translation_row_templates()
    return build_bundle()


if __name__ == "__main__":
    raise SystemExit(main())
